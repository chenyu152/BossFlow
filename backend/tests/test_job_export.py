import io
import unittest

from openpyxl import load_workbook

from backend.services.job_service import export_jobs_response


class JobExportTest(unittest.TestCase):
    def test_excel_export_includes_job_description(self):
        response = export_jobs_response([
            {
                "title": "Agent 应用开发工程师",
                "company": "示例公司",
                "city": "上海",
                "salary": "30-50K",
                "avg": 40,
                "exp": "3-5年",
                "edu": "本科",
                "cats": ["Agent"],
                "desc": "负责 Agent 应用开发、RAG 检索增强和服务落地。",
                "lastSeen": "2026-08-26",
                "url": "https://example.com/job_detail/demo.html",
            }
        ])

        workbook = load_workbook(io.BytesIO(response.body))
        sheet = workbook["岗位数据"]
        headers = [cell.value for cell in sheet[1]]
        description_column = headers.index("岗位详情") + 1

        self.assertEqual(sheet.cell(row=2, column=description_column).value, "负责 Agent 应用开发、RAG 检索增强和服务落地。")
        self.assertTrue(sheet.cell(row=2, column=description_column).alignment.wrap_text)
        self.assertEqual(sheet.column_dimensions["I"].width, 80)


if __name__ == "__main__":
    unittest.main()
