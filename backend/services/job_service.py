import io
import json
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import HTTPException
from fastapi.responses import Response

from backend.services.score_store import apply_scores_to_jobs


def _recruitment_observation_status(live_status: str, raw: str, checked_at: str) -> str:
    if not checked_at:
        return "not_checked"
    if live_status == "open":
        return "open_observed"
    if live_status == "closed":
        return "closed_observed"
    if raw == "login_required":
        return "login_required"
    if raw == "captcha_required":
        return "verification_required"
    if raw == "security_check":
        return "security_check"
    return "unknown_observed"


def _live_observation_fields(row: sqlite3.Row) -> dict[str, str]:
    live_status = row["live_status"] or ""
    raw = row["live_status_raw"] or ""
    checked_at = row["live_checked_at"] or ""
    return {
        "liveStatus": live_status,
        "liveStatusRaw": raw,
        "liveCheckedAt": checked_at,
        "liveClosedAt": row["live_closed_at"] or "",
        "liveCheckError": row["live_check_error"] or "",
        "recruitmentObservationStatus": _recruitment_observation_status(live_status, raw, checked_at),
        "recruitmentObservationRaw": raw,
        "recruitmentObservedAt": checked_at,
    }


def query_jobs(
    project_dir: Path,
    search: str = "",
    limit: int = 500,
    offset: int = 0,
    *,
    cities: Optional[list[str]] = None,
    tiers: Optional[list[str]] = None,
    categories: Optional[list[str]] = None,
    min_avg: float | None = None,
    max_avg: float | None = None,
    min_score: float | None = None,
    fit_levels: Optional[list[str]] = None,
    experience_risks: Optional[list[str]] = None,
    education_risks: Optional[list[str]] = None,
    recruitment_statuses: Optional[list[str]] = None,
    seen_since: str = "",
    scored_only: bool = False,
    sort_by: str = "salary_desc",
) -> Dict[str, Any]:
    db_path = project_dir / "jobs_data.db"
    if not db_path.exists() or db_path.stat().st_size == 0:
        return {"items": [], "total": 0}
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        clauses: list[str] = []
        params: List[Any] = []
        if search:
            clauses.append("(title LIKE ? OR company LIKE ? OR desc LIKE ?)")
            like = f"%{search}%"
            params.extend([like, like, like])
        for values, column in ((cities or [], "city"), (tiers or [], "tier")):
            normalized = [str(value).strip() for value in values if str(value).strip()]
            if normalized:
                clauses.append(f"{column} IN ({','.join('?' for _ in normalized)})")
                params.extend(normalized)
        if min_avg is not None:
            clauses.append("avg >= ?")
            params.append(float(min_avg))
        if max_avg is not None:
            clauses.append("avg <= ?")
            params.append(float(max_avg))
        if seen_since:
            clauses.append("last_seen >= ?")
            params.append(seen_since)
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        rows = conn.execute(
            f"""
            SELECT * FROM jobs {where}
            ORDER BY avg DESC, last_seen DESC, id DESC
            """,
            params,
        ).fetchall()
    finally:
        conn.close()

    items = []
    for row in rows:
        try:
            cats = json.loads(row["cats_json"] or "[]")
        except Exception:
            cats = []
        item = {
            "id": row["id"],
            "title": row["title"] or "",
            "company": row["company"] or "",
            "city": row["city"] or "",
            "salary": row["salary"] or "",
            "avg": float(row["avg"] or 0),
            "tier": row["tier"] or "",
            "exp": row["exp"] or "",
            "edu": row["edu"] or "",
            "cats": cats,
            "desc": row["desc"] or "",
            "url": row["url"] or "",
            "firstSeen": row["first_seen"] or "",
            "lastSeen": row["last_seen"] or "",
        }
        keys = set(row.keys())
        if "live_status" in keys:
            item.update(_live_observation_fields(row))
        items.append(item)
    items = apply_scores_to_jobs(project_dir.name, items)
    category_set = {str(value).strip() for value in (categories or []) if str(value).strip()}
    fit_set = {str(value).strip() for value in (fit_levels or []) if str(value).strip()}
    experience_set = {str(value).strip() for value in (experience_risks or []) if str(value).strip()}
    education_set = {str(value).strip() for value in (education_risks or []) if str(value).strip()}
    recruitment_set = {str(value).strip() for value in (recruitment_statuses or []) if str(value).strip()}
    filtered = []
    for item in items:
        if category_set and not category_set.intersection(str(value) for value in item.get("cats", [])):
            continue
        if min_score is not None and (item.get("score") is None or float(item["score"]) < float(min_score)):
            continue
        if scored_only and item.get("score") is None:
            continue
        if fit_set and str(item.get("fitLevel") or "") not in fit_set:
            continue
        if experience_set and str(item.get("experienceRisk") or "") not in experience_set:
            continue
        if education_set and str(item.get("educationRisk") or "") not in education_set:
            continue
        if recruitment_set and str(item.get("recruitmentObservationStatus") or "not_checked") not in recruitment_set:
            continue
        filtered.append(item)

    if sort_by == "newest":
        filtered.sort(key=lambda item: (str(item.get("lastSeen") or ""), int(item.get("id") or 0)), reverse=True)
    elif sort_by == "score_desc":
        filtered.sort(key=lambda item: (float(item.get("score") or -1), float(item.get("avg") or 0)), reverse=True)
    else:
        filtered.sort(key=lambda item: (float(item.get("avg") or 0), str(item.get("lastSeen") or ""), int(item.get("id") or 0)), reverse=True)
    total = len(filtered)
    return {"items": filtered[offset:offset + limit], "total": total, "offset": offset, "limit": limit}


def get_jobs_by_ids(project_dir: Path, job_ids: list[int]) -> list[dict[str, Any]]:
    ids = [int(job_id) for job_id in job_ids if int(job_id) > 0]
    if not ids:
        return []
    db_path = project_dir / "jobs_data.db"
    if not db_path.exists() or db_path.stat().st_size == 0:
        return []

    placeholders = ",".join("?" for _ in ids)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            f"SELECT * FROM jobs WHERE id IN ({placeholders}) ORDER BY avg DESC, last_seen DESC, id DESC",
            ids,
        ).fetchall()
    finally:
        conn.close()

    items: list[dict[str, Any]] = []
    for row in rows:
        try:
            cats = json.loads(row["cats_json"] or "[]")
        except Exception:
            cats = []
        item = {
            "id": row["id"],
            "title": row["title"] or "",
            "company": row["company"] or "",
            "city": row["city"] or "",
            "salary": row["salary"] or "",
            "avg": float(row["avg"] or 0),
            "tier": row["tier"] or "",
            "exp": row["exp"] or "",
            "edu": row["edu"] or "",
            "cats": cats,
            "desc": row["desc"] or "",
            "url": row["url"] or "",
            "firstSeen": row["first_seen"] or "",
            "lastSeen": row["last_seen"] or "",
        }
        keys = set(row.keys())
        if "live_status" in keys:
            item.update(_live_observation_fields(row))
        items.append(item)
    return apply_scores_to_jobs(project_dir.name, items)


def get_job_by_id(project_dir: Path, job_id: int) -> dict[str, Any]:
    jobs = get_jobs_by_ids(project_dir, [job_id])
    if not jobs:
        raise HTTPException(status_code=404, detail=f"Job not found: {project_dir.name}#{job_id}")
    return jobs[0]


def export_jobs_response(rows: list[dict[str, Any]]) -> Response:
    if not rows:
        raise HTTPException(status_code=404, detail="当前筛选条件下没有可导出的数据")
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Missing dependency: openpyxl") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "岗位数据"
    headers = ["岗位名称", "公司名称", "城市", "薪资", "平均薪资(K)", "经验", "学历", "分类", "最后活跃", "详情链接"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0067c0", end_color="0067c0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(
            [
                row["title"],
                row["company"],
                row["city"],
                row["salary"],
                row["avg"],
                row["exp"],
                row["edu"],
                ", ".join(row["cats"]),
                row["lastSeen"],
                row["url"],
            ]
        )
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 3
        ws.column_dimensions[col[0].column_letter].width = max(10, min(width, 42))
    out = io.BytesIO()
    wb.save(out)
    return Response(
        out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="boss_jobs_export.xlsx"'},
    )
